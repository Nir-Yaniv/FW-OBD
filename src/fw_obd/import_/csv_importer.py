"""Import device inventory from CSV or Excel (SolarWinds, PRTG, or generic exports)."""

from __future__ import annotations

import csv
import logging
import re
from dataclasses import dataclass
from io import StringIO
from pathlib import Path
from typing import Iterable, Optional

EXCEL_SUFFIXES = {".xlsx", ".xlsm"}

logger = logging.getLogger(__name__)

# Common column aliases from monitoring tools → canonical field names
COLUMN_ALIASES: dict[str, list[str]] = {
    "management_ip": [
        "site ip",
        "ip",
        "ip address",
        "management ip",
        "mgmt ip",
        "device ip",
        "node ip",
    ],
    "name": [
        "site name",
        "device name",
        "hostname",
        "caption",
        "name",
        "node name",
    ],
    "vendor": ["vendor", "brand", "manufacturer", "device type"],
    "location": ["location", "site", "city", "location city", "location site"],
    "region": ["region", "group", "custom property", "country", "location country"],
}


@dataclass
class ImportRow:
    name: str
    management_ip: str
    vendor: str = "Fortinet"
    location: str = ""
    region: str = ""


@dataclass
class ImportPreview:
    rows: list[ImportRow]
    skipped: int
    column_map: dict[str, str]


def _normalize_header(header: str) -> str:
    """Lowercase and treat underscores/hyphens as spaces so headers like
    ``IP_Address`` or ``Node-IP`` match aliases written with spaces."""
    cleaned = header.strip().lower().replace("_", " ").replace("-", " ")
    return " ".join(cleaned.split())


def _map_headers(fieldnames: list[str]) -> dict[str, str]:
    """Map CSV headers to canonical fields using alias table."""
    normalized = {_normalize_header(h): h for h in fieldnames}
    mapping: dict[str, str] = {}
    for canonical, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in normalized:
                mapping[canonical] = normalized[alias]
                break
    return mapping


_IPV4_RE = re.compile(r"^\d{1,3}(?:\.\d{1,3}){3}$")

# Header substrings that hint a column holds device names.
_NAME_HINTS = ("name", "caption", "host", "sysname", "node", "device", "firewall")


def _looks_like_ip(value: str) -> bool:
    v = value.strip()
    if _IPV4_RE.match(v):
        return all(0 <= int(part) <= 255 for part in v.split("."))
    # Loose IPv6: multiple colons and only hex/colon characters.
    return v.count(":") >= 2 and all(c in "0123456789abcdefABCDEF:" for c in v)


def _detect_ip_column(fieldnames: list[str], sample: list[dict]) -> Optional[str]:
    """Pick the column whose sample values most look like IP addresses."""
    best_col, best_score = None, 0.0
    for col in fieldnames:
        if not col:
            continue
        nonempty = [v for v in ((r.get(col) or "").strip() for r in sample) if v]
        if not nonempty:
            continue
        score = sum(_looks_like_ip(v) for v in nonempty) / len(nonempty)
        if score > best_score:
            best_col, best_score = col, score
    return best_col if best_score >= 0.5 else None


def _detect_name_column(
    fieldnames: list[str], sample: list[dict], ip_col: Optional[str]
) -> Optional[str]:
    """Pick a mostly-textual column to use as the device name."""
    candidates: list[str] = []
    for col in fieldnames:
        if not col or col == ip_col:
            continue
        nonempty = [v for v in ((r.get(col) or "").strip() for r in sample) if v]
        if not nonempty:
            continue
        texty = sum(
            not _looks_like_ip(v) and not v.replace(".", "").isdigit() for v in nonempty
        ) / len(nonempty)
        if texty >= 0.7:
            candidates.append(col)
    if not candidates:
        return None
    # Prefer a name-ish header; otherwise the first textual column.
    for col in candidates:
        if any(hint in _normalize_header(col) for hint in _NAME_HINTS):
            return col
    return candidates[0]


def _resolve_field_columns(
    fieldnames: list[str], sample: list[dict], column_map: Optional[dict[str, str]]
) -> dict[str, str]:
    """Resolve canonical field -> source header, with content-based fallback.

    1. Explicit column_map wins.
    2. Alias-based header matching (handles known SolarWinds/PRTG exports).
    3. Content sniffing for the required IP and name columns, so unfamiliar
       exports still import without any manual mapping.
    """
    if column_map:
        return column_map
    mapping = _map_headers(fieldnames)
    if "management_ip" not in mapping:
        detected = _detect_ip_column(fieldnames, sample)
        if detected:
            mapping["management_ip"] = detected
    if "name" not in mapping:
        detected = _detect_name_column(fieldnames, sample, mapping.get("management_ip"))
        if detected:
            mapping["name"] = detected
    return mapping


def _build_preview(
    fieldnames: list[str],
    dict_rows: Iterable[dict],
    column_map: Optional[dict[str, str]] = None,
) -> ImportPreview:
    """Build an ImportPreview from header names and dict-shaped rows.

    Shared by both the CSV and Excel parsers. ``column_map`` keys are canonical
    fields; values are exact source header names. When None, columns are resolved
    via alias matching plus content-based fallback (see _resolve_field_columns).
    """
    rows_data = list(dict_rows)
    effective_map = _resolve_field_columns(fieldnames, rows_data[:25], column_map)

    ip_col = effective_map.get("management_ip")
    name_col = effective_map.get("name")
    if not ip_col or not name_col:
        raise ValueError(
            "Could not automatically find management IP and device name columns. "
            f"Detected headers: {fieldnames}"
        )

    rows: list[ImportRow] = []
    skipped = 0
    for line in rows_data:
        ip = (line.get(ip_col) or "").strip()
        name = (line.get(name_col) or "").strip()
        if not ip or not name:
            skipped += 1
            continue
        vendor_col = effective_map.get("vendor")
        location_col = effective_map.get("location")
        region_col = effective_map.get("region")
        vendor = (line.get(vendor_col) or "Fortinet").strip() if vendor_col else "Fortinet"
        if vendor.lower() in ("fortigate", "fortinet"):
            vendor = "Fortinet"
        rows.append(
            ImportRow(
                name=name,
                management_ip=ip,
                vendor=vendor,
                location=(line.get(location_col) or "").strip() if location_col else "",
                region=(line.get(region_col) or "").strip() if region_col else "",
            )
        )

    return ImportPreview(rows=rows, skipped=skipped, column_map=effective_map)


def parse_csv_text(
    text: str,
    column_map: Optional[dict[str, str]] = None,
) -> ImportPreview:
    """Parse CSV content into ImportRow objects (see _build_preview for mapping)."""
    reader = csv.DictReader(StringIO(text))
    if not reader.fieldnames:
        return ImportPreview(rows=[], skipped=0, column_map={})
    return _build_preview(list(reader.fieldnames), reader, column_map)


def parse_csv_file(path: Path, column_map: Optional[dict[str, str]] = None) -> ImportPreview:
    text = path.read_text(encoding="utf-8-sig")
    return parse_csv_text(text, column_map=column_map)


def parse_excel_file(path: Path, column_map: Optional[dict[str, str]] = None) -> ImportPreview:
    """Parse the first sheet of an .xlsx/.xlsm workbook into ImportRow objects.

    Row 1 is treated as the header row. Cell values are coerced to stripped
    strings so numeric cells (e.g. an IP typed as a number) still map cleanly.
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        if ws is None:
            return ImportPreview(rows=[], skipped=0, column_map={})
        row_iter = ws.iter_rows(values_only=True)
        try:
            header = next(row_iter)
        except StopIteration:
            return ImportPreview(rows=[], skipped=0, column_map={})

        fieldnames = ["" if h is None else str(h).strip() for h in header]
        dict_rows: list[dict] = []
        for raw in row_iter:
            record = {
                fieldnames[i]: ("" if raw[i] is None else str(raw[i]))
                for i in range(len(fieldnames))
                if i < len(raw)
            }
            dict_rows.append(record)
    finally:
        wb.close()

    return _build_preview(fieldnames, dict_rows, column_map)


def parse_import_file(path: Path, column_map: Optional[dict[str, str]] = None) -> ImportPreview:
    """Dispatch to the CSV or Excel parser based on file extension."""
    if path.suffix.lower() in EXCEL_SUFFIXES:
        return parse_excel_file(path, column_map=column_map)
    return parse_csv_file(path, column_map=column_map)
